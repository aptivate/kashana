from __future__ import unicode_literals, absolute_import
import datetime

from django.db.models import F, Q

from collections import defaultdict

from django.views.generic import DetailView
from braces.views import LoginRequiredMixin
from spreadsheetresponsemixin.views import SpreadsheetResponseMixin
from logframe.api import build_period_filter
from logframe.models import (
    LogFrame, Result, Indicator, SubIndicator, Period, Milestone, Target,
    Actual
)
from logframe.period_utils import get_periods, get_period, periods_intersect
from bs4 import BeautifulSoup
from openpyxl.styles import Style, PatternFill, fills, Color, Font
from openpyxl.workbook.workbook import Workbook


GANTT_COL_WIDTH = 4
DATE_COL_WIDTH = 11
GANTT_COL_COLOR = '8FBC8F'


def solid_fill(color):
    return PatternFill(
        start_color=Color(color),
        end_color=Color(color),
        patternType=fills.FILL_SOLID)

gant_fill = solid_fill(GANTT_COL_COLOR)

default_style = {}
LEVEL_STYLES = defaultdict(lambda: default_style)
LEVEL_STYLES[1] = {'font': Font(color=Color('FFFFFF')), 'fill': solid_fill('505A6B')}  # impact
LEVEL_STYLES[2] = LEVEL_STYLES[1]  # outcome
LEVEL_STYLES[3] = LEVEL_STYLES[2]  # output
# LEVEL_STYLES[4] = {'fill': solid_fill('FFFF00')} # Period
LEVEL_STYLES[5] = {'fill': solid_fill('FFFF00')}  # Result
LEVEL_STYLES['activity'] = {'fill': solid_fill('FFFFCC')}  # Activity

LEVEL_STYLES['indicator'] = {'fill': solid_fill('FFF47F')}
LEVEL_STYLES['subindicator-header'] = {'fill': solid_fill('99CC99')}
LEVEL_STYLES['subindicator-name'] = {'fill': solid_fill('FFFFFF')}
LEVEL_STYLES['subindicator-milestone'] = {'fill': solid_fill('FFFFFF')}
LEVEL_STYLES['subindicator-total'] = {'fill': solid_fill('D0E1F4')}
LEVEL_STYLES['subindicator-rating'] = {'fill': solid_fill('FFCC99')}

RATING_STYLES = {
    'green': {'fill': solid_fill('99C473')},
    'yellow': {'fill': solid_fill('FFB900')},
    'red': {'font': Font(color=Color('FFFFFF')), 'fill': solid_fill('CA3C3C')},
    'lightest-grey': {'fill': solid_fill('FBFBFB')},
    'light-grey': {'fill': solid_fill('D6D6D6')},
    'grey': {'font': Font(color=Color('FFFFFF')), 'fill': solid_fill('666666')},
}


def html2txt(html):
    value = html if html else u""
    soup = BeautifulSoup(value, features='html5')
    return soup.get_text().strip()


class ExportObjects(object):
    def __init__(self, logframe):
        self.objects = self.get_objects(logframe)
        self.mapping = self.__map_objects(self.objects, self.parent_attr)

    def get_objects(self, logframe):
        filter_kwargs = {self.filter_arg: logframe}
        return self.model.objects.filter(**filter_kwargs)

    def __map_objects(self, objects, parent_attr):
        o_map = defaultdict(list)
        for item in objects:
            o_map[getattr(item, parent_attr)].append(item)
        return o_map


class ExportSubIndicator(ExportObjects):
    model = SubIndicator
    parent_attr = "indicator_id"
    filter_arg = "indicator__result__log_frame"

    def get_subindicator_targets(self, milestone):
        mapping = {}
        targets = Target.objects.filter(milestone=milestone).values("value", "subindicator")
        for target in targets:
            mapping[target['subindicator']] = target['value']
        return mapping

    def get_value(self, subindicator):
        values = Actual.objects.filter(
            subindicator=subindicator).order_by("-column__date")
        if values.exists():
            return values[0].value
        else:
            return ""

    def get_rating(self, subindicator):
        rating = subindicator.rating
        name = rating.name if rating else "Unrated"
        style = RATING_STYLES[rating.color] if rating else None
        rating_cell = {'value': name}

        if style:
            rating_cell['styles'] = style

        return rating_cell

    def render_head(self):
        return ['', 'Milestone for period', 'Total to date', 'Rating']

    def render(self, subindicator):
        return [
            {'value': subindicator.name, 'styles': LEVEL_STYLES['subindicator-name']},
            {'value': self.targets.get(subindicator.id, u""), 'styles': LEVEL_STYLES['subindicator-milestone']},
            {'value': self.get_value(subindicator), 'styles': LEVEL_STYLES['subindicator-total']},
            self.get_rating(subindicator),
        ]


class ExportIndicator(ExportObjects):
    model = Indicator
    parent_attr = "result_id"
    filter_arg = "result__log_frame"

    def render(self, indicator):
        return [
            indicator.name or u"",
            html2txt(indicator.description)
        ]


class LogframeDataMixin(object):

    # TODO: add to LogFrame model as method?
    def get_results(self, logframe):
        def get_children(parent):
            reordered.append(parent)
            for r in results:
                if r.parent_id == parent.id:
                    get_children(r)

        results = Result.objects.filter(log_frame=logframe)
        reordered = []
        for r in results:
            if not r.parent_id:
                get_children(r)
        return reordered

    def get_period(self, logframe):
        self.period = Period.objects.get(log_frame=logframe)

    def get_period_boundaries(self):
        period_start = datetime.date(self.year, self.period.start_month, 1)
        period_end = \
            datetime.date(self.year + 1, self.period.start_month, 1) - \
            datetime.timedelta(days=1)
        return (period_start, period_end)

    def get_activities(self, result):
        if not self.period:
            self.get_period(self.logframe)
        period_start, period_end = self.get_period_boundaries()

        invalid_range = {
            'start_date__gt': F('end_date'),
        }

        # period_filter from API already matches all of them except invalid
        activity_filter = build_period_filter(period_start, period_end,
                                              'start_date', 'end_date')

        return result.activities.filter(
            activity_filter | Q(**invalid_range)
        ).order_by('start_date').all()

    def add_activities(self, results):
        for result in results:
            # is last child.
            if not result.children.exists():
                yield (result, list(self.get_activities(result)))
            else:
                yield (result, None)

    @staticmethod
    def row_style(styles, cells):
        """ Apply a style to cells """
        cells_with_style = []

        for cell in cells:
            cells_with_style.append({
                'value': cell,
                'styles': styles,
            })

        return cells_with_style


class ExcelStyledWorkbookGenerator(object):
    # Overrides the one from SpreadsheetResponseMixin
    def generate_xlsx(self, data, headers=None, file=None):
        wb = Workbook()
        ws = wb.get_active_sheet()

        # Put in headers
        rowoffset = 0
        if headers:
            rowoffset = 1
            for c, headerval in enumerate(headers, 1):
                ws.cell(row=1, column=c).value = headerval

        # Put in data
        for r, row in enumerate(data, 1):
            for c, cellval in enumerate(row, 1):
                cell_style = None

                if isinstance(cellval, dict):
                    if 'styles' in cellval:
                        styles = cellval['styles']
                    else:
                        styles = None
                    cellval = cellval['value']
                    if styles:
                        cell_style = Style(**styles)

                cell = ws.cell(row=r + rowoffset, column=c)
                cell.value = cellval

                if cell_style:
                    cell.style = cell_style
        if file:
            wb.save(file)
        return wb


class ExportLogframeData(LoginRequiredMixin, LogframeDataMixin, ExcelStyledWorkbookGenerator, SpreadsheetResponseMixin, DetailView):
    model = LogFrame

    def get_milestone(self, logframe, end_date):
        '''
        If possible return first milestone after period or last one before
        it otherwise.
        '''
        milestones = Milestone.objects.filter(log_frame=logframe).order_by("date")
        if milestones.filter(date__gte=end_date).exists():
            milestone = milestones.filter(date__gte=end_date)[0]
        else:
            milestones = list(milestones)
            milestone = milestones[-1] if len(milestones) else None
        return milestone

    def get_export_head(self, period):
        start, end = period
        return ['Quarterly report', '', start.isoformat(), end.isoformat()]

    def add_row(self, row=None, indent_cells=0):
        existing = row if row else []
        self.data.append([u""] * indent_cells + existing)

    def get_data(self, **kwargs):
        '''
        Algorithm:
            fetch data
            print period header
            for each result:
                print result head
                for each indicator in result:
                    print indicator head
                    print subindicator head (milestone, total, RAG)
                    for each subindicator in indicator:
                        print subindicator (name, target, measurement, rating)
                    print empty line
                print empty line
        '''
        self.data = []
        logframe = self.get_object()
        self.get_period(logframe)
        period_interval = self.period.get_period(kwargs.get("period"))

        milestone = self.get_milestone(logframe, period_interval[1])

        results = self.get_results(logframe)
        ind = ExportIndicator(logframe)
        subind = ExportSubIndicator(logframe)
        subind.targets = subind.get_subindicator_targets(milestone)

        self.add_row(self.get_export_head(period_interval))
        self.add_row()

        for result in results:
            row = [
                result.name or u"",
                html2txt(result.description)
            ] + ["", ""]
            style = LEVEL_STYLES[result.level]
            row = self.row_style(style, row)
            self.add_row(row, 1)

            for indicator in ind.mapping[result.id]:
                self.add_row(
                    self.row_style(
                        LEVEL_STYLES['indicator'],
                        ind.render(indicator) + ["", ""]),
                    1)
                self.add_row(
                    self.row_style(
                        LEVEL_STYLES['subindicator-header'],
                        subind.render_head()
                    ), 1)
                for subindicator in subind.mapping[indicator.id]:
                    self.add_row(subind.render(subindicator), 1)
                self.add_row()
            self.add_row()
        return self.data

    # Overrides the one from SpreadsheetResponseMixin
    def render_setup(self, **kwargs):
        self.get_data(**self.kwargs)
        return [self.data, None]

    def get(self, request, *args, **kwargs):
        self.get_period(self.get_object())
        return self.render_excel_response()

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep',
          'Oct', 'Nov', 'Dec']


class ExportPlanMixin(ExcelStyledWorkbookGenerator):
    model = LogFrame

    @staticmethod
    def get_period_header(start, end=None, padding=0):
        months = MONTHS[start.month - 1:] + MONTHS[:start.month - 1]
        if end:
            # This works even if end is next year and end.month < start.month
            months = months[:end.month - start.month + 1]
        return [''] * padding + months

    @staticmethod
    def get_period_list(start, end):
        periods_start = get_periods(start, end, start.month, 12)
        periods = [get_period(period.isoformat(), 12) for period in periods_start]
        return periods

    @staticmethod
    def mark_row(start, end, periods, padding=0):
        marked = []
        marked_value = {
            'value': "",
            'styles': {'fill': gant_fill},
        }
        for p in periods:
            # print 'pp: ', p
            marked.append(marked_value if periods_intersect(p[0], p[1], start, end) else "")
        return [''] * padding + marked

    def add_row(self, row=None, indent_cells=0):
        existing = row if row else []
        self.data.append([u""] * indent_cells + existing)

    def get_data(self, **kwargs):
        self.data = []

        results = self.get_results(self.logframe)
        rows = self.add_activities(results)

        self.get_export_head()
        self.add_row()
        periods = self.get_plan_periods()

        for result, activities in rows:
            cols = [html2txt(result.name), html2txt(result.description)]
            style = LEVEL_STYLES[result.level]
            self.add_row(self.row_style(style, cols))

            if activities is not None:
                for activity in activities:
                    activity_cols = [
                        html2txt(activity.name),
                        html2txt(activity.description),
                        html2txt(activity.deliverables),
                        activity.start_date,
                        activity.end_date,
                    ]

                    style = LEVEL_STYLES['activity']
                    activity_cols = self.row_style(style, activity_cols)
                    activity_cols += self.mark_row(
                        activity.start_date, activity.end_date, periods)
                    self.add_row(activity_cols)

                self.add_row()  # Add blank row for every group

        return self.data

    # Overrides the one from SpreadsheetResponseMixin
    def render_setup(self, **kwargs):
        self.get_data(**self.kwargs)
        return [self.data, None]


class ExportAnnualPlan(LoginRequiredMixin, LogframeDataMixin, ExportPlanMixin,
                       SpreadsheetResponseMixin, DetailView):
    def format_worksheet(self, ws):
        for col in 'FGHIJKLMNOPQ':
            ws.column_dimensions[col].width = GANTT_COL_WIDTH

        ws.column_dimensions['D'].width = DATE_COL_WIDTH
        ws.column_dimensions['E'].width = DATE_COL_WIDTH

    def get_export_head(self):
        return [
            self.add_row(['%s Annual Report' % self.year]),
            self.add_row(['Name', 'Description', 'Deliverables', 'Start', 'End']
                         + self.get_period_header(self.start_date)),
        ]

    def get_plan_periods(self):
        return self.get_period_list(*get_period(self.start_date.isoformat(), 1))

    def get(self, request, *args, **kwargs):
        self.logframe = self.get_object()
        self.get_period(self.logframe)
        self.year = int(kwargs['year'])
        self.start_date = datetime.date(self.year, self.period.start_month, 1)

        self.filename = '%s_annual_plan.xlsx' % self.year

        return self.render_excel_response()


class ExportQuarterPlan(LoginRequiredMixin, LogframeDataMixin, ExportPlanMixin,
                        SpreadsheetResponseMixin, DetailView):
    def format_worksheet(self, ws):
        cols = 'FGHIJKLMNOPQ'
        period_length = 12 / self.period.num_periods
        for col in cols[:period_length]:
            ws.column_dimensions[col].width = GANTT_COL_WIDTH
        ws.column_dimensions['D'].width = DATE_COL_WIDTH
        ws.column_dimensions['E'].width = DATE_COL_WIDTH

    def get_export_head(self):
        title_date = "%s - %s %d" % (self.start_date.strftime("%B"),
                                     self.end_date.strftime("%B"),
                                     self.start_date.year)
        return [
            self.add_row(['%s Quarter Report' % title_date]),
            self.add_row(['Name', 'Description', 'Deliverables', 'Start', 'End']
                         + self.get_period_header(self.start_date, self.end_date)),
        ]

    def get_plan_periods(self):
        return self.get_period_list(*get_period(self.start_date.isoformat(),
                                                self.period.num_periods))

    def get_period_boundaries(self):
        return (self.start_date, self.end_date)

    def get_short_month(self, d):
        return d.strftime("%b").lower()

    def get(self, request, *args, **kwargs):
        self.logframe = self.get_object()
        self.get_period(self.logframe)
        self.year = int(kwargs['year'])
        month = int(kwargs['month'])

        start = datetime.date(self.year, month, 1)
        self.start_date, self.end_date = get_period(start.isoformat(), self.period.num_periods)

        self.filename = '%s-%s_%d_quarter_plan.xlsx' % (self.get_short_month(self.start_date),
                                                        self.get_short_month(self.end_date),
                                                        self.year)
        return self.render_excel_response()
