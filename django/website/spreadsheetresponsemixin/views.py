import types

from django.http import HttpResponse
from django.db.models.query import QuerySet
from openpyxl import Workbook
from StringIO import StringIO
import csv


class SpreadsheetResponseMixin(object):

    def render_excel_response(self, **kwargs):
        filename = self.get_filename(extension='xlsx')
        # Generate content
        self.data, self.headers = self.render_setup(**kwargs)
        # Setup response
        content_type = \
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response = HttpResponse(content_type=content_type)
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)
        # Add content and return response
        self.generate_xlsx(data=self.data, headers=self.headers, file=response)
        return response

    def render_csv_response(self, **kwargs):
        filename = self.get_filename(extension='csv')
        # Generate content
        self.data, self.headers = self.render_setup(**kwargs)
        # Build response
        content_type = 'text/csv'
        response = HttpResponse(content_type=content_type)
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)
        # Add content to response
        self.generate_csv(data=self.data, headers=self.headers, file=response)
        return response

    def render_setup(self, **kwargs):
        # Generate content
        queryset = kwargs.get('queryset')
        fields = self.get_fields(**kwargs)
        data = self.generate_data(queryset=queryset, fields=fields)
        headers = kwargs.get('headers')
        if not headers:
            headers = self.generate_headers(data, fields=fields)
        return data, headers

    def generate_data(self, queryset=None, fields=None):
        if not queryset:
            try:
                queryset = self.queryset
            except AttributeError:
                raise NotImplementedError(
                    "You must provide a queryset on the class or pass it in."
                )
        # After all that, have we got a proper queryset?
        assert isinstance(queryset, QuerySet)

        if fields:
            list_of_lists = queryset.values_list(*fields)
        else:
            list_of_lists = queryset.values_list()
        return list_of_lists

    def generate_headers(self, data, fields=None):
        model_fields = [field for field in data.model._meta.fields]
        model_field_dict = dict([(model.name, model)
                                for model in model_fields])
        if fields:
            model_fields = (model_field_dict[field] for field in fields)
        field_names = (field.verbose_name.title() for field in model_fields)
        return tuple(field_names)

    def format_worksheet(self, ws):
        """ Hook for custom formatting of the worksheet """
        pass

    def generate_xlsx(self, data, headers=None, file=None):
        from openpyxl.styles import Alignment
        from openpyxl.units import DEFAULT_ROW_HEIGHT
        from collections import defaultdict
        import textwrap

        MAX_WIDTH = 60

        wb = Workbook()
        ws = wb.get_active_sheet()

        # Put in headers
        rowoffset = 0
        if headers:
            rowoffset = 1
            for c, headerval in enumerate(headers, 1):
                ws.cell(row=1, column=c).value = headerval

        # fixed width
        # max_width, with wrap
        # auto width
        max_widths = defaultdict(int)

        # Put in data
        for r, row in enumerate(data, 1):
            row_heights = [-1]
            for c, cellval in enumerate(row, 1):
                cell = ws.cell(row=r + rowoffset, column=c)
                if type(cellval) == dict:
                    value = cellval['value']
                    if 'styles' in cellval:
                        cell.style = cell.style.copy(**cellval['styles'])
                    if 'width' in cellval:
                        max_widths[c] = cellval['width']
                else:
                    value = cellval

                if isinstance(value, types.StringTypes):
                    value_len = len(value)
                    if value_len > max_widths[c]:
                        max_widths[c] = value_len

                    if value_len > MAX_WIDTH:
                        cell.style = cell.style.copy(
                            alignment=Alignment(wrap_text=True))
                        height = len(textwrap.wrap(value, MAX_WIDTH))
                        row_heights.append(height)

                cell.value = value

            if row:
                ws.row_dimensions[r + rowoffset].height = max(row_heights) * DEFAULT_ROW_HEIGHT

        for c, max_width in max_widths.items():
            dimension = ws.column_dimensions.values()[c-1]
            width = max_widths[c]

            if width > MAX_WIDTH:
                dimension.width = MAX_WIDTH
            else:
                dimension.width = width

        self.format_worksheet(ws)
        if file:
            wb.save(file)
        return wb

    def generate_csv(self, data, headers=None, file=None):
        if not file:
            generated_csv = StringIO()
        else:
            generated_csv = file
        writer = csv.writer(generated_csv, dialect='excel')
        # Put in headers
        if headers:
            writer.writerow([unicode(s).encode('utf-8') for s in headers])

        # Put in data
        for row in data:
            writer.writerow([unicode(s).encode('utf-8') for s in row])
        return generated_csv

    def get_render_method(self, format):
        if format == 'excel':
            return self.render_excel_response
        elif format == 'csv':
            return self.render_csv_response
        raise NotImplementedError("Export format is not recognized.")

    def get_format(self, **kwargs):
        if 'format' in kwargs:
            return kwargs['format']
        elif hasattr(self, 'format'):
            return self.format
        raise NotImplementedError("Format is not defined.")

    def get_filename(self, **kwargs):
        if 'filename' in kwargs:
            return kwargs['filename']
        if hasattr(self, 'filename'):
            return self.filename
        default_filename = 'export'
        extension = kwargs.get('extension', 'out')
        return "{0}.{1}".format(default_filename, extension)

    def get_fields(self, **kwargs):
        if 'fields' in kwargs:
            return kwargs['fields']
        elif hasattr(self, 'fields') and self.fields is not None:
            return self.fields
        else:
            model = None
            if hasattr(self, 'model') and self.model is not None:
                model = self.model
            elif hasattr(self, 'queryset') and self.queryset is not None:
                model = self.queryset.model
            if model:
                return model._meta.get_all_field_names()
        return ()
